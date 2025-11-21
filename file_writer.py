import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string

from templates import COMPANY_INFO_LIST, COMPANY_CELL_TITLES, COMPANY_INFO_CELLS

def create_file(name: str, companies: str) -> None:
    """
    Creates an excel file using the given name as the filename and write desired headers.

    @param name a string containing the desired name for the excel file.
    @param companies a string containing the companies we are observing separated by commas.
    """

    wb = Workbook()
    ws = wb.active

    company_list = companies.split(',')

    ws['A1'] = 'Time'

    ws.merge_cells('B1:H1')
    ws.merge_cells('I1:O1')
    ws.merge_cells('P1:V1')
    ws.merge_cells('W1:AC1')
    ws.merge_cells('AD1:AJ1')
    ws.merge_cells('AK1:AQ1')
    ws.merge_cells('AR1:AX1')
    ws.merge_cells('AY1:BE1')
    ws.merge_cells('BF1:BL1')
    ws.merge_cells('BM1:BS1')
    ws.merge_cells('BT1:BZ1')

    ws.merge_cells('A1:A2')

    top_left_cell = ws['A1']
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in COMPANY_CELL_TITLES:
        ws[f'{cell}1'] = company_list[COMPANY_CELL_TITLES.index(cell)]

        top_left_cell = ws[f'{cell}1']
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for key in COMPANY_INFO_CELLS:
        for cell in COMPANY_INFO_CELLS[key]:
            ws[f'{cell}2'] = COMPANY_INFO_LIST[COMPANY_INFO_CELLS[key].index(cell)]

            top_left_cell = ws[f'{cell}2']
            top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the workbook
    wb.save(f"{name}.xlsx")

def update_file(name: str, companies: str, company_info: dict, time: str) -> None:
    """
    Checks the existance of the file. If it doesn't exist it is created, otherwise it is simply just updated
    to include all of the information from the company_info dictionary as well as the time.

    @param name the name of the excel file.
    @param companies the string containing the company tickers separated by commas.
    @param company_info the dict containing all of the info related to each company.
    @param time the timestamp string containing the hh:mm:ss.ms of the exact moment.
    """

    if not os.path.exists(f'{name}.xlsx'):
        create_file(name, companies)

    workbook = load_workbook(f'{name}.xlsx')
    ws = workbook.active

    empty_row = ws.max_row + 1

    ws.cell(row=empty_row, column=1, value=time)

    for index, company in enumerate(company_info):
        if company == 'Update Flag':
            continue

        for i, val in enumerate(company_info[company]):
            column_index = column_index_from_string(
                COMPANY_INFO_CELLS[index][i]
            )

            ws.cell(row=empty_row, column=column_index, value=company_info[company][val])

    workbook.save(f'{name}.xlsx')
    company_info['Update Flag'] = False
    print(f'Data successfully appended to {name}')

    return None
